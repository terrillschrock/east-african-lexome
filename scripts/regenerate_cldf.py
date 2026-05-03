"""
EALex CLDF Regeneration Script
================================
Reads the master spreadsheet (African_Lexome_CLDF_Template.xlsx) and writes
fresh CLDF-format files to the cldf/ folder of your repository.

USAGE
-----
1. Edit the two paths in CONFIGURATION below to match where your files live.
2. Run from the command line:
     Windows:  py regenerate_cldf.py
     Mac/Linux: python3 regenerate_cldf.py
3. After successful run, open GitHub Desktop, review the diff, commit, and push.

REQUIREMENTS
------------
  pip install openpyxl
  (or on Windows: py -m pip install openpyxl)
"""

import os, json, csv, sys

# ── CONFIGURATION ──────────────────────────────────────────────────────────────
# Edit these two paths to match your setup:

SPREADSHEET = os.path.expanduser(
    "~/Documents/Linguistics/East African Lexome/African_Lexome_CLDF_Template.xlsx"
)
REPO_CLDF_FOLDER = os.path.expanduser(
    "~/Documents/GitHub/East-African-Lexome/cldf"
)

# ──────────────────────────────────────────────────────────────────────────────

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl is not installed.")
    print("Install it with:  py -m pip install openpyxl")
    sys.exit(1)

if not os.path.exists(SPREADSHEET):
    print(f"ERROR: Spreadsheet not found at:\n  {SPREADSHEET}")
    print("Edit the SPREADSHEET path at the top of this script.")
    sys.exit(1)

if not os.path.exists(REPO_CLDF_FOLDER):
    print(f"ERROR: Repo cldf folder not found at:\n  {REPO_CLDF_FOLDER}")
    print("Edit the REPO_CLDF_FOLDER path at the top of this script.")
    sys.exit(1)

OUT = REPO_CLDF_FOLDER
print(f"Reading from: {SPREADSHEET}")
print(f"Writing to:   {OUT}\n")

wb = load_workbook(SPREADSHEET)

def sheet_to_rows(sheet_name):
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(cell is not None and str(cell).strip() != '' for cell in row):
            rows.append(dict(zip(headers, [('' if c is None else str(c)) for c in row])))
    return headers, rows

# ── languages.csv ─────────────────────────────────────────────────────────────
_, langs = sheet_to_rows("LanguageTable")
out_cols = ["ID","Name","Glottocode","ISO639P3code","Latitude","Longitude",
            "Macroarea","Family","Subfamily","Country","EA_Countries",
            "Speaker_Population","Endangerment","Source_Name","Comment"]
with open(f"{OUT}/languages.csv", "w", newline='', encoding='utf-8') as f:
    w = csv.DictWriter(f, fieldnames=out_cols)
    w.writeheader()
    for r in langs:
        out_row = {c: r.get(c, '') for c in out_cols}
        if not out_row.get('Macroarea'):
            out_row['Macroarea'] = 'Africa'
        w.writerow(out_row)
print(f"  languages.csv:    {len(langs):>5} rows")

# ── parameters.csv ────────────────────────────────────────────────────────────
_, params = sheet_to_rows("ParameterTable")
param_cols = ["ID","Name","CAWL_ID","Source","Tier","Concepticon_ID",
              "Concepticon_Gloss","Semantic_Field","Part_Of_Speech",
              "Swadesh_100","Swadesh_207","LJ_100","Comment"]
with open(f"{OUT}/parameters.csv", "w", newline='', encoding='utf-8') as f:
    w = csv.DictWriter(f, fieldnames=param_cols)
    w.writeheader()
    for r in params:
        w.writerow({c: r.get(c, '') for c in param_cols})
print(f"  parameters.csv:   {len(params):>5} rows")

# ── forms.csv ─────────────────────────────────────────────────────────────────
_, forms = sheet_to_rows("FormTable")
form_cols = ["ID","Language_ID","Parameter_ID","Form","Segments","Source_Form",
             "IPA","Orthography","Loan","LoanSource_Language","Cognate_ID",
             "Cognate_Set_ID","Value","Comment","Source","Data_Source_Type",
             "Source_Page","Elicitation_Context","Collector","Collection_Date",
             "Confidence"]
real_forms = [r for r in forms if r.get('ID') and r['ID'] not in ('form-001','example')]
with open(f"{OUT}/forms.csv", "w", newline='', encoding='utf-8') as f:
    w = csv.DictWriter(f, fieldnames=form_cols)
    w.writeheader()
    for r in real_forms:
        w.writerow({c: r.get(c, '') for c in form_cols})
print(f"  forms.csv:        {len(real_forms):>5} rows")

# ── cognates / cognatesets / orthographies ───────────────────────────────────
for sheet, fname, cols in [
    ("CognateTable", "cognates.csv",
     ["ID","Form_ID","Cognateset_ID","Alignment","Alignment_Method","Doubt","Comment","Source"]),
    ("CognatesetTable", "cognatesets.csv",
     ["ID","Description","Proto_Form","Proto_Language","Comment"]),
    ("OrthoRegistry", "orthographies.csv",
     ["Ortho_ID","Name","Language_Family","Region","Era","Colonial_Power",
      "Script","Tonal_Marking","Key_Features","IPA_Notes","Reference","Comment"]),
]:
    _, rows = sheet_to_rows(sheet)
    with open(f"{OUT}/{fname}", "w", newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        for r in rows:
            w.writerow({c: r.get(c, '') for c in cols})
    print(f"  {fname:<17} {len(rows):>5} rows")

# ── sources.bib ───────────────────────────────────────────────────────────────
_, sources = sheet_to_rows("SourceTable")
def bib_escape(s):
    if not s: return ""
    return s.replace("{","\\{").replace("}","\\}").replace("&","\\&")

bib_lines = []
for r in sources:
    sid = r.get('ID','').strip()
    if not sid: continue
    btype = r.get('Type','misc').strip() or 'misc'
    bib_lines.append(f"@{btype}{{{sid},")
    fields = []
    for csv_field, bib_field in [
        ('Author','author'), ('Year','year'), ('Title','title'),
        ('Publisher','publisher'), ('Place','address'),
        ('Journal','journal'), ('Volume','volume'), ('Pages','pages'),
        ('URL','url')
    ]:
        v = r.get(csv_field,'').strip()
        if v:
            fields.append(f"  {bib_field} = {{{bib_escape(v)}}}")
    note_parts = []
    for f in ['Language_Covered','Region_Covered','Source_Type','Comment']:
        v = r.get(f,'').strip()
        if v:
            note_parts.append(f"{f}: {v}")
    if note_parts:
        fields.append(f"  note = {{{bib_escape(' | '.join(note_parts))}}}")
    bib_lines.append(",\n".join(fields))
    bib_lines.append("}\n")

with open(f"{OUT}/sources.bib", "w", encoding='utf-8') as f:
    f.write("\n".join(bib_lines))
print(f"  sources.bib:      {len(sources):>5} entries")

print("\nDone! Now open GitHub Desktop, review the diff, commit, and push.")
